from datetime import datetime
from abc import ABC, abstractmethod
from enum import Enum
import uuid
from typing import Literal
import hashlib
import os
import mimetypes
import csv
from openpyxl import Workbook

class Notification():
    def __init__(self, id: str, message: str, receipients: str, created_at=None): # receipients include classroom name
        self.id = id
        self.message = message
        self.receipients = receipients
        self.created_at = created_at or datetime.now().isoformat()
        self.is_read = False

    def to_dict(self):
        return {
            "Type": "Notification",
            "ID": self.id,
            "Message": self.message,
            "Receipient_class": self.receipients,
            "Created_at": self.created_at
        }

    def send(self):
        count_s = 0
        count_p = 0
        for student in STUDENTS.values():
            if student.grade == self.receipients:
                receipient = STUDENTS[student._id]
                receipient._notifications.append(self)
                receipient.notifications.append(self.message)
                count_s += 1
                receipient_p = PARENTS[student.parent]
                receipient_p._notifications.append(self)
                receipient_p.notifications.append(self.message)
                count_p += 1
        print(f"Xabar {count_s} o'quvchi va {count_p} ota-onalarga yuborildi.")

    def mark_as_read(self):
        self.is_read = True

class Assignment:
    def __init__(self, id: int, title: str, description: str, deadline: str, subject: str, teacher_id: int, class_id: str, difficulty: str):
        self.id = id
        self.title = title
        self.description = description
        self.deadline = deadline
        self.subject = subject
        self.teacher_id = teacher_id
        self.class_id = class_id
        self.difficulty = difficulty

        self.submissions: dict[int, str] = {} # {student id: content}
        self.grades: dict[int, Grade] = {} # {student id: assignment grade}

        ASSIGNMENTS[self.id] = self # adding to the dictionary

    def to_dict(self):
        return {
            "Type": "Assignment",
            "ID": self.id,
            "Title": self.title,
            "Description": self.description,
            "Deadline": self.deadline,
            "Subject": self.subject,
            "Teacher_ID": self.teacher_id,
            "Class_ID": self.class_id,
            "Difficulty": self.difficulty,
            "Submissions": self.submissions,
            "Grades": self.grades
        }
    
    def add_submissions(self, student_id: int, content: str): # CONTENT
        if student_id in self.submissions:
            print(f"Student {student_id} vazifani allaqachon topshirgan. Ustidan yozilmaydi.")
            return
        self.submissions[student_id] = content # content path
        print("Vazifa {student_id} tomonidan qabul qilindi.")

    def check_submission(self, student_id):
        if student_id in self.submissions.keys():
            size_in_bytes = os.path.getsize(self.submissions[student_id]) # content size
            size_in_kb = size_in_bytes / 1024
            print(f"File is {size_in_kb:.2f} KB")
            mime_type, _ = mimetypes.guess_type(self.submissions[student_id])
            print("File type type:", mime_type)


    def set_grade(self, student_id: int, grade: float):
        if student_id in self.submissions:
            if grade >= 0 and grade <= 100:
                self.grades[student_id] = grade

                if grade <= 100 and grade >= 85: grade = 5
                elif grade < 85 and grade <= 75: grade = 4
                elif grade < 75 and grade <= 60: grade = 3
                elif grade < 60 and grade >= 0: grade = 2
                else: raise ValueError("Ball 0 - 100 shkalada emas.")
                grade_id = str(uuid.uuid4())
                GRADES[grade_id] = Grade(grade_id, student_id, self.subject, grade, self.deadline, self.teacher_id)
                print(f"{student_id}ga {grade} ball qo'yildi.")
            else:
                raise ValueError("Ball 0 - 100 shkalada emas.")
        else:
            print(f"{student_id}ning topshirgan vazifasi topilmadi.")

    def get_status(self):
        now = datetime.now().isocalendar()
        if now < self.deadline:
            print("Vazifalar qabul qilinmoqda.")
        print("Dedlayn tugadi. Vazifalar qabul qilinmaydi.")

class Grade():
    def __init__(self, id: int, student_id: int, subject: str, value: int, date: str, teacher_id: int):
        self.id = id
        self.student_id = student_id
        self.subject = subject
        self.value = value
        student = STUDENTS.get(student_id)
        if student:
            student.grades.setdefault(subject, []).append(self)
        self.date = date
        self.teacher_id = teacher_id

    def to_dict(self):
        return {
            "Type": "Grade",
            "ID": self.id,
            "Student_ID": self.student_id,
            "Subject": self.subject,
            "Value": self.value,
            "Date": self.date,
            "Teacher_ID": self.teacher_id
            }

    def update_grade(self, value: Literal[1, 2, 3, 4, 5]):
        self.value = value

    def get_grade_info(self):
        info = {
            "Student_ID" : self.student_id,
            "Subject" : self.subject,
            "Value" : self.value,
            "Date" : self.date,
            "Teacher_ID" : self.teacher_id
        }
        for k, v in info.items():
            print(f"{k.replace("_", " ")}: {v}")
        return info

class Schedule:
    def __init__(self, id: int, class_id: str, day: str):
        self.id = id
        self.class_id = class_id
        self.day = day
        self.lessons : dict[str, dict[str, int]] = {} # lessons to be added later in form {time: {subject : teacher_id}}
        SCHEDULES[self.id] = self

    def to_dict(self):
        return {
            "Type": "Schedule",
            "ID": self.id,
            "Grade" : self.class_id,
            "Day": self.day,
            "Lessons": self.lessons
        }

    def add_lesson(self, time: str, subject: str, teacher_id: int): # time indicates the starting time of the lesson.
        if time not in self.lessons:
            for sche in SCHEDULES.values():
                if sche.day == self.day:
                    for k, v in sche.lessons.items():
                        teacher_in_sche = list(v.values())
                        if k == time and teacher_in_sche[0] == teacher_id:
                            raise ValueError("O'qituvchining bu vaqtda darsi bor.")
            self.lessons[time] = {subject : teacher_id}
            teacher = TEACHERS[teacher_id]
            if self.class_id not in teacher.classes:
                teacher.classes.append(self.class_id)
            if subject not in teacher.subjects:
                teacher.subjects.append(subject)

        else:
            response = input(f"Bu vaqtda {self.lessons[time]} darslari bor. Ustidan yozish kerakmi? (1-ha)\n")
            if response == 1:
                self.lessons[time] = {subject : teacher_id}

    def view_schedule(self):
        schedule = {
            "Schedule_ID" : self.id,
            "Class_ID" : self.class_id,
            "Day" : self.day,
            "Lessons": self.lessons
        }
        for k, v in schedule.items():
            print(f"{k.replace("_", " ")} : {v}")
        return schedule

    def remove_lesson(self, time):
        if time in self.lessons:
            lesson = self.lessons.pop(time)
            subject = list(lesson)[0]
            print(f"Dars jadvalidan {subject} darsi o'chirildi.")
        else:
            print(f"Dars jadvalida {time} vaqt oralig'ida dars mavjud emas.")





class User_role(Enum): #For User
    ADMIN = "Admin"
    TEACHER = "Teacher"
    STUDENT = "Student"
    PARENT = "Parent"

class AbstractRole(ABC):
    def __init__(self, _id: int, _full_name: str, _email: str, _password: str, _created_at: str = None):
        self._id = _id
        self._full_name = _full_name
        self._email = _email
        self._password_hush = self.password_hush_func(_password)
        self._created_at = _created_at or datetime.now().isoformat()

    def password_hush_func(self, password):
        return hashlib.sha256(password.encode()).hexdigest()


    @abstractmethod
    def get_profile(self):
        pass

    @abstractmethod
    def update_profile(self):
        pass

class User(AbstractRole):
    def __init__(self, _id: int, _full_name: str, _email: str, _password_hash: str, 
                phone: str, address: str, role: User_role, _notifications: list[Notification] = None,_created_at: str = None):
        super().__init__(_id, _full_name, _email, _password_hash, _created_at)
        self.role = role
        self.phone = phone
        self.address = address
        if _notifications is None:
            self._notifications = []
        else:
            self._notifications = _notifications
        self.notifications : list[str] = []
        for note in self._notifications:
            self.notifications.append(note.message)

    def get_profile(self):
        return {
            "id": self._id,
            "full name": self._full_name,
            "email": self._email,
            "phone" : self.phone,
            "address" : self.address,
            "role": self.role.value,
            "created at": self._created_at
        }

    def update_profile(self, full_name = None, email = None, password_hash = None):
        if full_name:
            self._full_name = full_name
        if email:
            self._email = email
        if password_hash:
            self._password_hush = password_hash

    def add_notification(self, notification_message: str):
        notification = {
            "id": uuid.uuid4, # every notification should have a unique id
            "message": notification_message
        }
        self._notifications.append(notification)

    def view_notifications(self):
        for n in range(len(self._notifications)):
            print(self._notifications[n].message)
            # response = input("Xabar o'qildimi? (1 - ha)\n")
            # if response == 1:
                # del self._notifications[n]

    def delete_notification(self, id: str):
        self._notifications = [i for i in self._notifications if i["id"] != id]

class Student(User):
    def __init__(self, _id: int, _full_name: str, _email: str, _password_hash: str,  
                grade: str, _parent: int = None, phone: str=None, address: str=None, _notifications: list[Notification] = None, _created_at: str = None):
        super().__init__(_id, _full_name, _email, _password_hash, phone, address, User_role.STUDENT, _notifications, _created_at)
        self.grade = grade
        self.subjects: dict[str, int] = {} # {Subject Name: Teacher ID}
        self.assignments: dict[str, str] = {} # {Assignment ID: status}
        self.grades: dict[str, list[float]] = {} # {Subject: grades}
        self.parent = _parent
        self.notifications : list[str] = []
        for note in self._notifications:
            self.notifications.append(note.message)

    def to_dict(self):
        return {
            "Role": "Student",
            "ID": self._id,
            "Full_name": self._full_name,
            "Email": self._email,
            "Password_hash": self._password_hush,
            "Grade": self.grade,
            "Parent": self.parent,
            "Phone": self.phone,
            "Address": self.address,
            "Subjects": self.subjects,
            "Assignments": self.assignments,
            "Grades": self.grades,
            "Notifications" : self.notifications,
            "Created_at": self._created_at
        }

    def submit_assignment(self, assignment_id, content):
        self.assignments[assignment_id] = 'submitted'
        assignment = ASSIGNMENTS.get(assignment_id)
        assignment.submissions[self._id] = content
        return f"Vazifa {assignment_id} muvaffaqiyatli qabul qilindi."
        # WHERE TO STORE THE CONTENT?

    def view_grades(self, subject=None):
        if subject:
            return self.grades.get(subject, [])
        return self.grades

    def calculate_average_grade(self):
        total = 0
        count = 0
        for grade_list in self.grades.values():
            total += sum(grade_list)
            count += len(grade_list)
        if count == 0:
            return 0.0
        return round(total / count, 2)
    
    def add_parent(self, parent):
        self.parent = parent

    def view_schedule(self):
        classes : list[Schedule] = [] # holds all the six days of the calendar for the student's class
        week : list[Schedule] = [None, None, None, None, None, None]
        for schedule in SCHEDULES.values():
            if schedule.class_id == self.grade:
                classes.append(schedule)
        for a_class in classes:
            if a_class.day == 'Monday':
                week.insert(0, a_class)
            elif a_class.day == "Tuesday":
                week.insert(1, a_class)
            elif a_class.day == "Wednesday":
                week.insert(2, a_class)
            elif a_class.day == "Thursday":
                week.insert(3, a_class)
            elif a_class.day == "Friday":
                week.insert(4, a_class)
            elif a_class.day == "Saturday":
                week.insert(5, a_class)
            else:
                pass
        for day in week:
            if day:
                print(f"\n{day.day}\n-----------------")
                for k, v in day.lessons.items():
                    print(f"{k} - {v}")
                    

class Teacher(User):
    def __init__(self, _id: int, _full_name: str, _email: str, _password_hash: str, subjects: list[str],
                phone: str = None, address: str = None, _notifications: list[Notification] = None,_created_at: str = None):
        super().__init__(_id, _full_name, _email, _password_hash, phone, address, User_role.TEACHER, _notifications, _created_at)
        self.subjects : list[str] = subjects # [subject name 1, subject name 2, ...]
        self.classes: list[str] = [] # [class id 1, class id 2, ...]
        # assignments are stored in a global dictionary
        self.notifications : list[str] = []
        for note in self._notifications:
            self.notifications.append(note.message)

    def to_dict(self):
        return {
            "Role": "Teacher",
            "ID": self._id,
            "Full_Name": self._full_name,
            "Email": self._email,
            "Password_hash": self._password_hush,
            "Phone_number": self.phone,
            "Address": self.address,
            "Notifications": self.notifications,
            "Created_at": self._created_at,
            "Subjects" : self.subjects,
            "Classes" : self.classes
        }
        
    def create_assignment(self, title: str, description: str, deadline: str, subject: str, class_id: str, difficulty: str):
        assignment_id = len(ASSIGNMENTS) + 1
        assignment = Assignment(assignment_id, title, description, deadline, subject, self._id, class_id, difficulty)
        ASSIGNMENTS[assignment_id] = assignment
        notification_id = str(uuid.uuid4())
        NOTIFICATIONS[notification_id] = Notification(notification_id, "Yangi vazifa!", class_id)
        NOTIFICATIONS[notification_id].send()
        print("Vazifa muvaffaqiyatli qo'shildi.")
    
    def grade_assignment(self, assignment_id: int, student_id: int, grade: float):
        assignment = ASSIGNMENTS.get(assignment_id)
        if assignment and student_id in assignment.submissions.keys():
            assignment.set_grade(student_id, grade)
            print(f"O'quvchi {student_id}ga {grade} ball qo'yildi.")
        elif assignment:
            print(f"O'quvchi {student_id} vazifani yuklamagan.")
        else:
            print("Vazifa topilmadi.")
        
    def view_student_progress(self, student_id):
        student = STUDENTS.get(student_id)
        if student:
            progress = {
                subj: grades
                for subj, grades in student.grades.items()
                if subj in self.subjects
            }
            return progress
        raise ValueError(f"O'quvchi {student_id} topilmadi.")

class Parent(User):
    def __init__(self, _id: int, _full_name: str, _email: str, _password_hash: str, 
                children: list[int], phone: str=None, address: str=None, _notifications: list[Notification] = None,_created_at: str = None):
        super().__init__(_id, _full_name, _email, _password_hash, phone, address, User_role.PARENT, _notifications, _created_at)
        self.children = children # list of student ids
        self.notifications : list[str] = []
        for note in self._notifications:
            self.notifications.append(note.message)

        for child in self.children:
            if child in STUDENTS.keys():
                STUDENTS[child].parent = _id
            else:
                raise ValueError(f"O'quvchi {child} ro'yxatda yo'q.")
            
    def to_dict(self):
        return {
            "Role": "Parent",
            "ID" : self._id,
            "Full_name": self._full_name,
            "Email":self._email,
            "Password_hash": self._password_hush,
            "Children": self.children,
            "Phone": self.phone,
            "Address": self.address,
            "Notifications": self.notifications,
            "Created_at": self._created_at
        }

    def view_child_grades(self, child_id: int):
        child = STUDENTS.get(child_id)
        if child:
            if child_id not in self.children:
                print(f"Sizda o'quvchi {child_id}ning ma'lumotlarini ko'rishga ruxsat yo'q.")
                return
            return child.grades
        else:
            raise ValueError(f"O'quvchi {child_id} topilmadi.")
    
    def view_child_assignments(self, child_id: int):
        child = STUDENTS.get(child_id)
        if child:
            if child_id not in self.children:
                print(f"Sizda o'quvchi {child_id}ning ma'lumotlarini ko'rishga ruxsat yo'q.")
                return
            return child.assignments
        else:
            raise ValueError(f"O'quvchi {child_id} topilmadi.")

    def receive_child_notifications(self, child_id: int):
        child = STUDENTS.get(child_id)
        if child:
            if child_id not in self.children:
                print(f"Sizda o'quvchi {child_id}ning baholarini ko'rishga ruxsat yo'q.")
                return
            for n in range(len(self._notifications)):
                print(self._notifications[n].message)
                # is_read = int(input("Xabar o'qildimi? (1 - ha)\n"))
                # if is_read == 1:
                    # self._notifications.pop(n) # So the old notifications are deleted regularly.
        else:
            raise ValueError(f"O'quvchi {child_id} topilmadi.")

class Admin(User):
    def __init__(self,  _id: int, _full_name: str, _email: str, _password_hash: str, 
                permissions: list[str], phone: str=None, address: str=None, _notifications: list[Notification] = None,_created_at: str = None):
        super().__init__(_id, _full_name, _email, _password_hash, phone, address, User_role.PARENT , _notifications, _created_at)
        self.permissions = permissions # list of permissions
        self.notifications : list[str] = []
        for note in self._notifications:
            self.notifications.append(note.message)

    def to_dict(self):
        return {
            'Role' : 'Admin',
            'ID' : self._id,
            'Full_name' : self._full_name,
            'Email' : self._email,
            'Password_Hash' : self._password_hush,
            'Permissions' : self.permissions,
            'Phone_number' : self.phone,
            'Address' : self.address,
            'Notifications' : self.notifications,
            'Created_at' : self._created_at
        }
    
    def prepare_export_tables(self):
        return {
            "ADMINS": [admin.to_dict() for admin in ADMINS.values()],
            "STUDENTS": [s.to_dict() for s in STUDENTS.values()],
            "TEACHERS": [t.to_dict() for t in TEACHERS.values()],
            "PARENTS": [p.to_dict() for p in PARENTS.values()],
            "ASSIGNMENTS": [a.to_dict() for a in ASSIGNMENTS.values()],
            "GRADES": [g.to_dict() for g in GRADES.values()],
            "SCHEDULES": [s.to_dict() for s in SCHEDULES.values()],
            "NOTIFICATIONS": [n.to_dict() for n in NOTIFICATIONS.values()],
            }
    
    def export_to_csv(self):
        tables = self.prepare_export_tables()
        os.makedirs("csv_exports", exist_ok=True)
        for name, table in tables.items():
            if table:
                with open(f"csv_exports/{name}.csv", "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=table[0].keys())
                    writer.writeheader()
                    writer.writerows(table)

    def export_to_xlsx(self):
        tables = self.prepare_export_tables()
        from openpyxl import Workbook

        def clean_value(val):
            if isinstance(val, (list, dict)):
                return str(val)
            return val

        wb = Workbook()
        first_sheet_created = False

        for name, table in tables.items():
            if not table:
                continue

            ws = wb.create_sheet(title=name)
            ws.append(list(table[0].keys()))
            for row in table:
                cleaned_row = [clean_value(val) for val in row.values()]
                ws.append(cleaned_row)
            first_sheet_created = True

        if "Sheet" in wb.sheetnames and first_sheet_created:
            del wb["Sheet"]

        wb.save("data_export.xlsx")

    def export_to_sql(self):
        tables = self.prepare_export_tables()
        def sql_safe(val):
            if val is None:
                return "NULL"
            elif isinstance(val, str):
                # Escape single quotes by doubling them
                return f"'{val.replace('\'', '\'\'')}'"
            elif isinstance(val, (list, dict)):
                return f"'{str(val).replace('\'', '\'\'')}'"  # still wrap as string
            else:
                return str(val)
            
        with open("data_export.sql", "w", encoding="utf-8") as f:
            for name, table in tables.items():
                if not table:
                    continue
                columns = table[0].keys()
                create_stmt = f"CREATE TABLE {name} (\n" + ",\n".join(f"  {col} NVARCHAR(MAX)" for col in columns) + "\n);\n"
                f.write(create_stmt)

                for row in table:
                    values = ", ".join(sql_safe(v) for v in row.values())
                    insert_stmt = f"INSERT INTO {name} ({', '.join(columns)}) VALUES ({values});\n"
                    f.write(insert_stmt)
                f.write("\n")

    
    def add_user(self, user: User): # A PERENT CLASS, MIGHT CLASH
        USERS[user._id] = user
        match user.role.name:
            case "TEACHER":
                TEACHERS[user._id] = user
            case "STUDENT":
                STUDENTS[user._id] = user
            case "PARENT":
                PARENTS[user._id] = user
            case "ADMIN":
                ADMINS[user._id] = user

    def remove_user(self, user_id):
        if user_id in USERS:
            match USERS[user_id].role.name:
                case "TEACHER":
                    del TEACHERS[user_id]
                case "STUDENT":
                    del STUDENTS[user_id]
                case "PARENT":
                    del PARENTS[user_id]
            del USERS[user_id]
            print(f"Foydalanuvchi {user_id} bazadan o'chirildi.")
        else:
            print(f"Foydalanuvchi {user_id} bazada topilmadi.")
    
    def generate_report(self):
        report = {
            "Total_users" : len(USERS),
            "Total_students" : len(STUDENTS),
            "Total_teachers" : len(TEACHERS),
            "Total_parents" : len(PARENTS),
            "Total_assignments" : len(ASSIGNMENTS),
            "Total_grades" : len(GRADES)
        }
        for k, v in report.items():
            print(f"{k.replace("_", " ")}: {v}")
        return report




        


USERS: dict[int, User] = {}
ADMINS: dict[int, Admin] = {}
STUDENTS: dict[int, Student] = {}
TEACHERS: dict[int, Teacher] = {}
PARENTS: dict[int, Parent] = {}
ASSIGNMENTS: dict[int, Assignment] = {}
GRADES: dict[str, Grade] = {}
SCHEDULES: dict[int, Schedule] = {}
NOTIFICATIONS: dict[str, Notification] = {}